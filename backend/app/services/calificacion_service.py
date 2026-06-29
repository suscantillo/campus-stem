import io
import uuid
from collections import defaultdict

from fastapi import HTTPException, status
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.db.models.calificaciones import AsignacionJuez, Calificacion, CriterioRubrica, PuntajeCriterio
from app.db.models.platform_controls import SINGLETON_ID, PlatformControls
from app.db.models.teams import Equipo
from app.db.models.users import RolUsuario, Usuario
from app.schemas.calificacion import (
    AsignacionToggleRequest,
    CalificacionRequest,
    CalificacionResponse,
    CalificacionStatusResponse,
    CriterioCreate,
    CriterioResponse,
    CriterioUpdate,
    EquipoAsignadoResponse,
    EquipoInfo,
    JuezInfo,
    MatrizAsignacionesResponse,
    NotaJuezResponse,
    OrdenRequest,
    PuntajeResponse,
    ResultadoEquipoResponse,
    ResultadosResponse,
)


class CalificacionService:
    def __init__(self, db: AsyncSession):
        self.db = db

    # ── helpers ────────────────────────────────────────────────────────────────

    async def _require_calificacion_cerrada(self) -> None:
        controls = await self.db.get(PlatformControls, SINGLETON_ID)
        if controls and controls.calificacion_abierta:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="No se puede modificar la rúbrica con la calificación abierta.",
            )

    async def _require_calificacion_abierta(self) -> None:
        controls = await self.db.get(PlatformControls, SINGLETON_ID)
        if not controls or not controls.calificacion_abierta:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="La fase de calificación no está abierta.",
            )

    def _build_calificacion_response(self, cal: Calificacion) -> CalificacionResponse:
        puntajes = [
            PuntajeResponse(
                criterio_id=p.criterio_id,
                criterio_nombre=p.criterio.nombre if p.criterio else None,
                puntaje=p.puntaje,
                puntaje_maximo=p.criterio.puntaje_maximo if p.criterio else None,
            )
            for p in cal.puntajes
        ]
        return CalificacionResponse(
            id=cal.id,
            juez_id=cal.juez_id,
            equipo_id=cal.equipo_id,
            comentario=cal.comentario,
            total=sum(p.puntaje for p in cal.puntajes),
            created_at=cal.created_at,
            puntajes=puntajes,
        )

    # ── Criterios ──────────────────────────────────────────────────────────────

    async def list_criterios(self) -> list[CriterioResponse]:
        result = await self.db.execute(
            select(CriterioRubrica).order_by(CriterioRubrica.orden, CriterioRubrica.created_at)
        )
        return [CriterioResponse.model_validate(c) for c in result.scalars().all()]

    async def create_criterio(self, data: CriterioCreate) -> CriterioResponse:
        await self._require_calificacion_cerrada()
        criterio = CriterioRubrica(**data.model_dump())
        self.db.add(criterio)
        await self.db.commit()
        await self.db.refresh(criterio)
        return CriterioResponse.model_validate(criterio)

    async def update_criterio(self, criterio_id: uuid.UUID, data: CriterioUpdate) -> CriterioResponse:
        await self._require_calificacion_cerrada()
        criterio = await self.db.get(CriterioRubrica, criterio_id)
        if not criterio:
            raise HTTPException(status_code=404, detail="Criterio no encontrado.")
        for field, value in data.model_dump(exclude_none=True).items():
            setattr(criterio, field, value)
        await self.db.commit()
        await self.db.refresh(criterio)
        return CriterioResponse.model_validate(criterio)

    async def delete_criterio(self, criterio_id: uuid.UUID) -> None:
        await self._require_calificacion_cerrada()
        criterio = await self.db.get(CriterioRubrica, criterio_id)
        if not criterio:
            raise HTTPException(status_code=404, detail="Criterio no encontrado.")
        await self.db.delete(criterio)
        await self.db.commit()

    async def reorder_criterios(self, data: OrdenRequest) -> list[CriterioResponse]:
        await self._require_calificacion_cerrada()
        for orden, criterio_id in enumerate(data.ids):
            criterio = await self.db.get(CriterioRubrica, criterio_id)
            if criterio:
                criterio.orden = orden
        await self.db.commit()
        return await self.list_criterios()

    # ── Asignaciones ───────────────────────────────────────────────────────────

    async def get_matriz(self) -> MatrizAsignacionesResponse:
        jueces_res = await self.db.execute(
            select(Usuario).where(Usuario.rol == RolUsuario.JUEZ).order_by(Usuario.nombre_completo)
        )
        jueces = jueces_res.scalars().all()

        equipos_res = await self.db.execute(select(Equipo).order_by(Equipo.nombre))
        equipos = equipos_res.scalars().all()

        asign_res = await self.db.execute(select(AsignacionJuez))
        asignaciones = asign_res.scalars().all()

        asignados = [f"{a.juez_id}:{a.equipo_id}" for a in asignaciones]

        return MatrizAsignacionesResponse(
            jueces=[JuezInfo(id=j.id, nombre_completo=j.nombre_completo) for j in jueces],
            equipos=[EquipoInfo(id=e.id, nombre=e.nombre) for e in equipos],
            asignados=asignados,
        )

    async def reset_asignaciones(self) -> MatrizAsignacionesResponse:
        jueces_res = await self.db.execute(
            select(Usuario).where(Usuario.rol == RolUsuario.JUEZ)
        )
        jueces = jueces_res.scalars().all()
        equipos_res = await self.db.execute(select(Equipo))
        equipos = equipos_res.scalars().all()

        await self.db.execute(delete(AsignacionJuez))

        for juez in jueces:
            for equipo in equipos:
                self.db.add(AsignacionJuez(juez_id=juez.id, equipo_id=equipo.id))

        await self.db.commit()
        return await self.get_matriz()

    async def toggle_asignacion(self, data: AsignacionToggleRequest) -> MatrizAsignacionesResponse:
        result = await self.db.execute(
            select(AsignacionJuez).where(
                AsignacionJuez.juez_id == data.juez_id,
                AsignacionJuez.equipo_id == data.equipo_id,
            )
        )
        existing = result.scalar_one_or_none()

        if existing:
            # Verificar que no haya calificación enviada
            cal_res = await self.db.execute(
                select(Calificacion).where(
                    Calificacion.juez_id == data.juez_id,
                    Calificacion.equipo_id == data.equipo_id,
                )
            )
            if cal_res.scalar_one_or_none():
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="No se puede eliminar: ya existe una calificación enviada.",
                )
            await self.db.delete(existing)
        else:
            self.db.add(AsignacionJuez(juez_id=data.juez_id, equipo_id=data.equipo_id))

        await self.db.commit()
        return await self.get_matriz()

    # ── Toggle calificación ────────────────────────────────────────────────────

    async def toggle_calificacion(self, abierta: bool) -> CalificacionStatusResponse:
        if abierta:
            # Validar que haya criterios
            count_res = await self.db.execute(select(func.count()).select_from(CriterioRubrica))
            count = count_res.scalar_one()
            if count == 0:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="Define al menos un criterio en la rúbrica antes de abrir la calificación.",
                )
            # Crear asignaciones por defecto si no existen
            asign_count_res = await self.db.execute(select(func.count()).select_from(AsignacionJuez))
            if asign_count_res.scalar_one() == 0:
                await self.reset_asignaciones()

        controls = await self.db.get(PlatformControls, SINGLETON_ID)
        if not controls:
            raise HTTPException(status_code=500, detail="Platform controls not initialized.")
        controls.calificacion_abierta = abierta
        await self.db.commit()

        return await self.get_status()

    async def get_status(self) -> CalificacionStatusResponse:
        controls = await self.db.get(PlatformControls, SINGLETON_ID)
        abierta = controls.calificacion_abierta if controls else False

        jueces_res = await self.db.execute(
            select(func.count()).select_from(Usuario).where(Usuario.rol == RolUsuario.JUEZ)
        )
        total_jueces = jueces_res.scalar_one()

        equipos_res = await self.db.execute(select(func.count()).select_from(Equipo))
        total_equipos = equipos_res.scalar_one()

        enviadas_res = await self.db.execute(select(func.count()).select_from(Calificacion))
        enviadas = enviadas_res.scalar_one()

        asign_res = await self.db.execute(select(func.count()).select_from(AsignacionJuez))
        esperadas = asign_res.scalar_one()

        return CalificacionStatusResponse(
            calificacion_abierta=abierta,
            total_jueces=total_jueces,
            total_equipos=total_equipos,
            calificaciones_enviadas=enviadas,
            calificaciones_esperadas=esperadas,
        )

    # ── Juez: scoring ──────────────────────────────────────────────────────────

    async def get_mis_equipos(self, juez_id: uuid.UUID) -> list[EquipoAsignadoResponse]:
        asign_res = await self.db.execute(
            select(AsignacionJuez)
            .where(AsignacionJuez.juez_id == juez_id)
            .options(selectinload(AsignacionJuez.equipo))
        )
        asignaciones = asign_res.scalars().all()

        cal_res = await self.db.execute(
            select(Calificacion)
            .where(Calificacion.juez_id == juez_id)
            .options(selectinload(Calificacion.puntajes).selectinload(PuntajeCriterio.criterio))
        )
        cals_by_equipo = {c.equipo_id: c for c in cal_res.scalars().all()}

        items = []
        for a in asignaciones:
            if a.equipo is None:
                continue
            cal = cals_by_equipo.get(a.equipo_id)
            items.append(
                EquipoAsignadoResponse(
                    equipo_id=a.equipo_id,
                    equipo_nombre=a.equipo.nombre,
                    calificacion=self._build_calificacion_response(cal) if cal else None,
                )
            )
        items.sort(key=lambda x: x.equipo_nombre)
        return items

    async def enviar_calificacion(
        self, juez_id: uuid.UUID, equipo_id: uuid.UUID, data: CalificacionRequest
    ) -> CalificacionResponse:
        await self._require_calificacion_abierta()

        # Verificar asignación
        asign_res = await self.db.execute(
            select(AsignacionJuez).where(
                AsignacionJuez.juez_id == juez_id,
                AsignacionJuez.equipo_id == equipo_id,
            )
        )
        if not asign_res.scalar_one_or_none():
            raise HTTPException(status_code=403, detail="No tienes asignado este equipo.")

        # Verificar que no haya enviado ya
        existing_res = await self.db.execute(
            select(Calificacion).where(
                Calificacion.juez_id == juez_id,
                Calificacion.equipo_id == equipo_id,
            )
        )
        if existing_res.scalar_one_or_none():
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Ya enviaste tu calificación para este equipo.",
            )

        # Validar puntajes contra criterios
        criterios_res = await self.db.execute(select(CriterioRubrica))
        criterios = {c.id: c for c in criterios_res.scalars().all()}

        for p in data.puntajes:
            criterio = criterios.get(p.criterio_id)
            if not criterio:
                raise HTTPException(status_code=422, detail=f"Criterio {p.criterio_id} no existe.")
            if p.puntaje > criterio.puntaje_maximo:
                raise HTTPException(
                    status_code=422,
                    detail=f"Puntaje {p.puntaje} excede el máximo de {criterio.puntaje_maximo} para '{criterio.nombre}'.",
                )

        cal = Calificacion(
            juez_id=juez_id,
            equipo_id=equipo_id,
            comentario=data.comentario,
        )
        self.db.add(cal)
        await self.db.flush()

        for p in data.puntajes:
            self.db.add(PuntajeCriterio(
                calificacion_id=cal.id,
                criterio_id=p.criterio_id,
                puntaje=p.puntaje,
            ))

        await self.db.commit()

        # Reload con relaciones
        cal_res = await self.db.execute(
            select(Calificacion)
            .where(Calificacion.id == cal.id)
            .options(selectinload(Calificacion.puntajes).selectinload(PuntajeCriterio.criterio))
        )
        cal = cal_res.scalar_one()
        return self._build_calificacion_response(cal)

    # ── Resultados ─────────────────────────────────────────────────────────────

    async def get_resultados(self) -> ResultadosResponse:
        criterios = await self.list_criterios()

        jueces_res = await self.db.execute(
            select(Usuario).where(Usuario.rol == RolUsuario.JUEZ).order_by(Usuario.nombre_completo)
        )
        jueces = jueces_res.scalars().all()

        equipos_res = await self.db.execute(select(Equipo).order_by(Equipo.nombre))
        equipos = equipos_res.scalars().all()

        asign_res = await self.db.execute(select(AsignacionJuez))
        # juez_id -> set of equipo_ids
        asign_by_juez: dict[uuid.UUID, set[uuid.UUID]] = defaultdict(set)
        # equipo_id -> set of juez_ids
        asign_by_equipo: dict[uuid.UUID, set[uuid.UUID]] = defaultdict(set)
        for a in asign_res.scalars().all():
            asign_by_juez[a.juez_id].add(a.equipo_id)
            asign_by_equipo[a.equipo_id].add(a.juez_id)

        cal_res = await self.db.execute(
            select(Calificacion).options(selectinload(Calificacion.puntajes))
        )
        # (juez_id, equipo_id) -> total
        cal_totals: dict[tuple, int] = {}
        for cal in cal_res.scalars().all():
            total = sum(p.puntaje for p in cal.puntajes)
            cal_totals[(cal.juez_id, cal.equipo_id)] = total

        resultados = []
        for equipo in equipos:
            jueces_asignados = asign_by_equipo.get(equipo.id, set())
            notas = []
            totals_recibidos = []
            for juez in jueces:
                if juez.id not in jueces_asignados:
                    continue
                total = cal_totals.get((juez.id, equipo.id))
                notas.append(NotaJuezResponse(
                    juez_id=juez.id,
                    juez_nombre=juez.nombre_completo,
                    total=total,
                ))
                if total is not None:
                    totals_recibidos.append(total)

            recibidas = len(totals_recibidos)
            esperadas = len(jueces_asignados)
            promedio = (sum(totals_recibidos) / recibidas) if recibidas > 0 else None

            if recibidas == 0:
                estado = "Sin calificar"
            elif recibidas < esperadas:
                estado = "Parcial"
            else:
                estado = "Completo"

            resultados.append(ResultadoEquipoResponse(
                posicion=0,
                equipo_id=equipo.id,
                equipo_nombre=equipo.nombre,
                notas=notas,
                promedio=promedio,
                calificaciones_recibidas=recibidas,
                calificaciones_esperadas=esperadas,
                estado=estado,
            ))

        # Ordenar: promedio desc (None al final), luego nombre asc
        resultados.sort(key=lambda r: (r.promedio is None, -(r.promedio or 0), r.equipo_nombre))
        for i, r in enumerate(resultados):
            r.posicion = i + 1

        return ResultadosResponse(
            criterios=criterios,
            resultados=resultados,
            jueces=[JuezInfo(id=j.id, nombre_completo=j.nombre_completo) for j in jueces],
        )

    async def export_excel(self) -> bytes:
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise HTTPException(
                status_code=500,
                detail="openpyxl no instalado. Ejecuta: pip install openpyxl",
            ) from exc

        data = await self.get_resultados()

        wb = openpyxl.Workbook()

        # ── Hoja 1: Resultados ───────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Resultados"

        header_fill = PatternFill("solid", fgColor="012854")
        header_font = Font(bold=True, color="FFFFFF")
        bold = Font(bold=True)

        headers = ["Pos.", "Equipo", "Promedio", "Estado", "Cal. recibidas"]
        for j in data.jueces:
            headers.append(j.nombre_completo)

        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for r in data.resultados:
            row = [
                r.posicion,
                r.equipo_nombre,
                round(r.promedio, 2) if r.promedio is not None else "—",
                r.estado,
                f"{r.calificaciones_recibidas}/{r.calificaciones_esperadas}",
            ]
            # nota por juez (en orden de data.jueces)
            notas_by_juez = {n.juez_id: n.total for n in r.notas}
            for j in data.jueces:
                t = notas_by_juez.get(j.id)
                row.append(t if t is not None else "—")
            ws1.append(row)

        for col in ws1.columns:
            ws1.column_dimensions[col[0].column_letter].width = 18

        # ── Hoja 2: Detalle ──────────────────────────────────────────────────
        ws2 = wb.create_sheet("Detalle")

        detail_headers = ["Equipo", "Juez"] + [c.nombre for c in data.criterios] + ["Total"]
        for col, h in enumerate(detail_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Re-fetch calificaciones con puntajes para detalle
        cal_res = await self.db.execute(
            select(Calificacion).options(selectinload(Calificacion.puntajes))
        )
        cals = cal_res.scalars().all()
        cal_map: dict[tuple, Calificacion] = {(c.juez_id, c.equipo_id): c for c in cals}

        criterio_ids = [uuid.UUID(str(c.id)) for c in data.criterios]
        equipos_map = {r.equipo_id: r.equipo_nombre for r in data.resultados}
        jueces_map = {j.id: j.nombre_completo for j in data.jueces}

        row_idx = 2
        for r in data.resultados:
            for nota in r.notas:
                cal = cal_map.get((nota.juez_id, r.equipo_id))
                if cal is None:
                    continue
                puntajes_by_criterio = {p.criterio_id: p.puntaje for p in cal.puntajes}
                detail_row = [
                    equipos_map.get(r.equipo_id, "—"),
                    jueces_map.get(nota.juez_id, "—"),
                ]
                for cid in criterio_ids:
                    detail_row.append(puntajes_by_criterio.get(cid, "—"))
                detail_row.append(nota.total if nota.total is not None else "—")
                ws2.append(detail_row)
                row_idx += 1

        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
